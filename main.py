import time
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook

driver_path = '/usr/local/bin/chromedriver'
url = 'https://hdeli.by/app/randomajzer/'

a = 0


try:
    wb = load_workbook("output2.xlsx")
except FileNotFoundError:
    wb = Workbook()

ws = wb.active

# Определение начальной строки для записи данных
row_number = 1

while a != 100:
    try:
        driver = webdriver.Chrome()
        driver.get(url)
        selenium_button = driver.find_element(By.CSS_SELECTOR, 'button')
        count_input = driver.find_element(By.ID, 'count')
        # Очистить текущее значение в поле input и ввести новое значение "2"
        count_input.clear()
        count_input.send_keys("2")
        selenium_button.click()
        time.sleep(3)

        # Получение HTML-кода страницы после динамической генерации контента
        page_source = driver.page_source

        soup = BeautifulSoup(page_source, 'html.parser')

        div_elements = soup.find_all('div', class_='blog-title d-block text-dark')
        h6 = []
        p = []
        for div_element in div_elements:

            h6.append([h6_element.text for h6_element in div_element.find_all('h6')])
            p.append([p_element.text for p_element in div_element.find_all('p')])

        # Записываем значения в ячейки
        h6_values = ', '.join([item for sublist in h6 for item in sublist])
        p_values = ', '.join([item for sublist in p for item in sublist])

        # Записать строки в ячейки
        ws.cell(row=row_number, column=1, value=h6_values)
        ws.cell(row=row_number, column=2, value=p_values)

        # Увеличиваем номер строки для следующей записи данных
        row_number += 1

    except Exception as e:
        print("Ошибка:", e)

    finally:
        # Закрытие браузера
        driver.quit()
        a += 1

# Сохраняем книгу Excel
wb.save("output2.xlsx")